import pandas as pd
import os, re, json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image

# from services.appscan_parser import AppScanParser

class AppScanExcelReport:
    def __init__(self):
        self.template_path = os.path.join(os.getcwd(), "data/report_templates", "template_web.xlsx")
        # self.template_path = os.path.join("D:/GitHub/report", "data/report_templates", "template_web.xlsx")
        self.severity_order = ['嚴重', '高', '中', '低', '參考資訊']
        
    def _df_log(self, index, df):
        log_dict = {
            "index":index+1, 
            "web_url": df["information"].loc[0, "web_url"], 
            "start_time": df["information"].loc[0, "start_time"]
            }
        df_log = pd.DataFrame(log_dict, index=[0])
        return df_log
    
    def _cause_count_total(self, df_risk_web, need_canse):
        cause_count_total = pd.DataFrame({"severity": self.severity_order, "amount": [0] * len(self.severity_order)})
        cause_count = df_risk_web.groupby("severity").size().reset_index(name='amount')

        cause_count_total = pd.concat([cause_count_total, cause_count]).groupby("severity", as_index=False)["amount"].sum()
        cause_count_total["severity"] = pd.Categorical(cause_count_total["severity"], categories=self.severity_order, ordered=True)
        cause_count_total = cause_count_total.sort_values("severity").reset_index(drop=True)
        # 留嚴重、高、中
        cause_count_total = cause_count_total.iloc[:need_canse]   # 留嚴重、高、中
        return cause_count_total
        
    def _write_dataframe_to_sheet(self, ws, df, start_row=3):
        pattern = re.compile(r"{{\s*(\w+)\s*}}")
        placeholder_cells = {}
        
        for cell in ws[start_row - 1]:
            if cell.value and isinstance(cell.value, str):
                match = pattern.fullmatch(cell.value)
                if match:
                    placeholder_cells[match.group(1)] = cell.column
        
        df = df[list(placeholder_cells.keys())]
        
        for r_idx, row in enumerate(df.itertuples(index=False), start=start_row):
            for col_name, col_idx in placeholder_cells.items():
                cell = ws.cell(row=r_idx, column=col_idx, value=getattr(row, col_name))
                # 讀取模板的格式（避免 `StyleProxy` 錯誤）
                template_cell = ws.cell(row=start_row - 1, column=col_idx)
                # 設定文字換行  ##改成中間左邊
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                # **設置字體**
                cell.font = Font(
                    name=template_cell.font.name,
                    size=template_cell.font.size,
                    bold=template_cell.font.bold,
                    italic=template_cell.font.italic,
                    underline=template_cell.font.underline,
                    color=template_cell.font.color
                )
                # **設置填充（背景顏色）**
                cell.fill = PatternFill(
                    fill_type=template_cell.fill.fill_type,
                    start_color=template_cell.fill.start_color,
                    end_color=template_cell.fill.end_color
                )
                # **複製邊框（格線）**
                cell.border = Border(
                    left=Side(style=template_cell.border.left.style, color=template_cell.border.left.color),
                    right=Side(style=template_cell.border.right.style, color=template_cell.border.right.color),
                    top=Side(style=template_cell.border.top.style, color=template_cell.border.top.color),
                    bottom=Side(style=template_cell.border.bottom.style, color=template_cell.border.bottom.color),
                )
        
        ws.delete_rows(start_row - 1)
    
    def _write_count_img_to_sheet(self, ws, df, img_path, need_canse):  
        total = pd.DataFrame({'severity': '總計', 'amount': df.amount.sum()}, index=[0])
        df = pd.concat([df, total])
        self._write_dataframe_to_sheet(ws, df)
        
        img_palce = f"A{need_canse+4}"
        img = Image(img_path)
        img.width = 400  # 設置圖片寬度
        img.height = 260  # 設置圖片高度
        ws.add_image(img, img_palce)  # 插入圖片

    def _excel(self, task_folder_path, json_paths, need_canse, img_name):
        output_path = os.path.join(task_folder_path, f"{img_name}.xlsx")
        # 整理成需要的df
        df_logs = pd.DataFrame(columns=['index', 'web_url', 'start_time'])
        df_risk_webs = pd.DataFrame(columns=['name', 'goal_url', 'severity', 'cvss', 'cvss_vector', 
                                            'cve', 'url', 'entity', 'risk', 'cause', 'solution', 
                                            'mark', 'solution_description', 'solution_suggest'])
        for index, json_path in enumerate(json_paths):
            dfs = self._to_df(json_path)
            df_log = self._df_log(index, dfs)
            df_logs = pd.concat([df_logs, df_log])
            df_risk_web = dfs["risk_web"]
            # 留需要的風險程度
            df_risk_web = df_risk_web[df_risk_web["severity"].isin(self.severity_order[:need_canse])]
            df_risk_webs = pd.concat([df_risk_webs, df_risk_web])

        wb = load_workbook(self.template_path)        
        # 檢測日誌
        ws_01 = wb.worksheets[0]
        self._write_dataframe_to_sheet(ws_01, df_logs, start_row=3)
        # 弱點統計
        ws_02 = wb.worksheets[1]
        cause_count_total = self._cause_count_total(df_risk_webs, need_canse)
        img_path = os.path.join(task_folder_path, "image", f"{img_name}.jpg")
        self._write_count_img_to_sheet(ws_02, cause_count_total, img_path, need_canse)
        # 詳細資料
        ws_03 = wb.worksheets[2]
        self._write_dataframe_to_sheet(ws_03, df_risk_webs, start_row=3)
        wb.save(output_path)
        print(f"Excel 報告已成功生成：{output_path}")

    def generate_excel_report(self, task_folder_path, report_type, need_canse=3):
        print(task_folder_path)
        json_paths = []
        file_names = []
        for root, _, files in os.walk(task_folder_path):
            json_paths.extend([os.path.join(root, f) for f in files if f.endswith('.json')])
            file_names.extend([os.path.splitext(f)[0] for f in files if f.endswith('.json')])

        if report_type == "single":
            img_name = os.path.basename(os.path.dirname(task_folder_path))
            self._excel(task_folder_path, json_paths, need_canse, img_name)
        elif report_type == "multiple":
            for jj, ff in zip(json_paths, file_names):
                jj = [jj]
                self._excel(task_folder_path, jj, need_canse, ff)

    @staticmethod
    def _to_df(json_path):
        # 讀取 JSON 檔案
        with open(json_path, "r", encoding="utf-8") as f:
            json_data = json.load(f)

        # 轉回原本的 dict[dict[DataFrame]]
        dfs = {key: pd.DataFrame(records) for key, records in json_data.items()}
        return dfs


def test():
    # appscan_files = [
    #     "Appscan_https___ginandjuice.shop_.scan",
    #     "Appscan_http___testasp.vulnweb.com_.scan"
    # ]
    task_folder_path = r"D:\GitHub\report\data\uploadproject\MyTaskName6\appscan"
    
    # parser = AppScanParser()
    # parser.run("C:/Program Files (x86)/HCL/AppScan Standard/AppScanCMD.exe", appscan_files, task_folder_path, test=True)
    
    need_canse=3  # 只需要幾個風險填幾個
    report_type = "single"
    report_generator_excel = AppScanExcelReport()
    report_generator_excel.generate_excel_report(task_folder_path, report_type, need_canse)

if __name__ == "__main__":
    test()
