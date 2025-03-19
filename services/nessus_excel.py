import os, json
import pandas as pd
import re

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

from services.nessus_trans import ApplinaceTranslate
from services.nessus_parser import NessusParser

class NessusExcelGenerater:
    def __init__(self):
        print("NessusExcelGenerater --- loaded")

        # 模板位置目前是寫死的(未來可以考慮多個版型可選)
        self.template_path = os.path.join(os.getcwd(), "data/report_templates", "template_va.xlsx")
        self.severity_num_to_ch = {
            0: "參考資訊",
            1: "低",
            2: "中",
            3: "高",
            4: "嚴重",
        }
        self.severity_order = ["參考資訊", "低", "中", "高", "嚴重"]
    
    # 以保留格式的方式將 DataFrame 寫入 Excel 同時清除佔位符
    def write_dataframe_to_sheet(self, ws, df, start_row=3):
        ### c_name => ip system os pluginid severity pluginname description solution port protocal cvss3 cvss2

        # 找到模板中的佔位符
        placeholder_cells = {}
        pattern = re.compile(r"{{\s*(\w+)\s*}}")  # 匹配 {{ xxx }}

        for cell in ws[start_row - 1]:  # 讀取表頭行，尋找佔位符
            if cell.value and isinstance(cell.value, str):
                match = pattern.fullmatch(cell.value)
                if match:
                    col_name = match.group(1)  # 提取佔位符名稱
                    placeholder_cells[col_name] = cell.column  # 記錄佔位符對應的列號

        # 確保 DataFrame 的欄位名稱與佔位符匹配
        df = df[list(placeholder_cells.keys())]  # 過濾 df，只保留模板中存在的欄位
        
        # 寫入數據，保持格式
        for r_idx, row in enumerate(df.itertuples(index=False), start=start_row):
            for col_name, col_idx in placeholder_cells.items():
                cell = ws.cell(row=r_idx, column=col_idx, value=getattr(row, col_name))

                # 讀取模板的格式（避免 `StyleProxy` 錯誤）
                template_cell = ws.cell(row=start_row - 1, column=col_idx)

                # **設置字體**
                cell.font = Font(
                    name=template_cell.font.name,
                    size=template_cell.font.size,
                    bold=template_cell.font.bold,
                    italic=template_cell.font.italic,
                    underline=template_cell.font.underline,
                    color=template_cell.font.color
                )

                # **設置填充（背景顏色）**
                cell.fill = PatternFill(
                    fill_type=template_cell.fill.fill_type,
                    start_color=template_cell.fill.start_color,
                    end_color=template_cell.fill.end_color
                )

                # **複製邊框（格線）**
                cell.border = Border(
                    left=Side(style=template_cell.border.left.style, color=template_cell.border.left.color),
                    right=Side(style=template_cell.border.right.style, color=template_cell.border.right.color),
                    top=Side(style=template_cell.border.top.style, color=template_cell.border.top.color),
                    bottom=Side(style=template_cell.border.bottom.style, color=template_cell.border.bottom.color),
                )

                # **設置對齊方式（置中、靠左、靠右）**
                cell.alignment = Alignment(
                    horizontal=template_cell.alignment.horizontal,  # 左對齊/置中/右對齊
                    vertical=template_cell.alignment.vertical,      # 頂對齊/置中/底對齊
                    # wrap_text=template_cell.alignment.wrap_text,    # 是否自動換行
                    # shrink_to_fit=template_cell.alignment.shrink_to_fit,  # 是否自動縮小字體以適應單元格
                    # indent=template_cell.alignment.indent          # 縮排
                )
        
        # 刪除佔位符行（避免留存 `{name}` `{age}` 等）
        ws.delete_rows(start_row - 1)

    def sheet_2(self, ws, df, project_path):
        table = df.groupby("severity").size().reset_index(name="count")
        placeholders = {
            "{{ crit }}": str(table[table["severity"] == "4"]["count"].sum()),
            "{{ high }}": str(table[table["severity"] == "3"]["count"].sum()),
            "{{ med }}": str(table[table["severity"] == "2"]["count"].sum()),
            "{{ low }}": str(table[table["severity"] == "1"]["count"].sum()),
            "{{ total }}": str(df.shape[0]),
        }
        
        host_table = df.groupby("severity")["ip"].nunique().reset_index(name="count")
        host_placeholders = {
            "{{ host_crit }}": str(host_table[host_table["severity"] == "4"]["count"].sum()),
            "{{ host_high }}": str(host_table[host_table["severity"] == "3"]["count"].sum()),
            "{{ host_med }}": str(host_table[host_table["severity"] == "2"]["count"].sum()),
            "{{ host_low }}": str(host_table[host_table["severity"] == "1"]["count"].sum()),
        }

        # 遍歷 Excel 的所有儲存格，尋找佔位符並替換
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    # 替換數據佔位符
                    if cell.value in placeholders:
                        cell.value = placeholders[cell.value]  # 替換數值
                        cell.font = Font(bold=True)  # 設定為加粗
                    elif cell.value in host_placeholders:
                        cell.value = host_placeholders[cell.value]  # 替換數值
                        cell.font = Font(bold=True)  # 設定為加粗
                    # 處理 image_ 開頭的佔位符
                    elif cell.value.startswith("{{ image_"):
                        image_name = cell.value[9:-3] + ".jpg"  # 取得圖片檔名
                        image_path = os.path.join(project_path, image_name)
                        
                        try:
                            pil_img = PILImage.open(image_path)
                            aspect_ratio = pil_img.width / pil_img.height
                            new_width = 450
                            new_height = int(new_width / aspect_ratio)  # 保持比例

                            img = Image(image_path)
                            img.width = new_width
                            img.height = new_height
                            ws.add_image(img, cell.coordinate)  # 插入圖片到對應位置
                            
                            cell.value = ""  # 清空原本的佔位符
                        
                        except Exception as e:
                            print(f"圖片 {image_path} 無法插入，錯誤訊息：{e}")
        
        

    def translate_row(self, row):
        try:
            trans = ApplinaceTranslate()
            trans_data = trans.trans_run(row["pluginid"])  # 查詢 pluginid
            row["pluginName"] = trans_data["pluginName"]
            row["description"] = trans_data["description"]
            row["solution"] = trans_data["solution"]
            return row
        
        except Exception as e:
            print(f"翻譯過程中出現錯誤(in generateexceldata.py): {e}")

    def generate_report(self, project_path):
        """將 Nessus 解析結果輸出至 Excel 並保持格式，刪除佔位符"""
        output_path = os.path.join(project_path, "outputfile.xlsx")

        information = self._to_df(os.path.join(project_path, "nessus.json"))
        df, df_scan_info = information["df_nessus"], information["df_nessus_scan_info"]
 
        wb = load_workbook(self.template_path)

        ws_01 = wb.worksheets[0]  # 第1個工作表(檢測日誌)
        self.write_dataframe_to_sheet(ws_01, df_scan_info, start_row=3)

        ws_02 = wb.worksheets[1]  # 第2個工作表(弱點統計)
        df_02 = df[df['severity'] != "0"].copy()
        self.sheet_2(ws_02, df_02, project_path)
        ## 各主機弱點數分佈表
        df_host = df_02.pivot_table(index='ip', columns='severity', aggfunc='size', fill_value=0)
        df_host.columns = ["ip_low", "ip_med", "ip_high", "ip_crit"]
        df_host['ip_total'] = df_host.sum(axis=1)
        df_host = df_host.rename_axis(columns=None).reset_index()
        self.write_dataframe_to_sheet(ws_02, df_host, start_row=4)
        
        
        ## 轉換 severity 欄位的數字 -> 文字
        df["severity"] = df["severity"].astype(int).map(self.severity_num_to_ch).fillna(df["severity"])
        # trans plugin to mandarin
        df = df.apply(self.translate_row, axis=1) 
        df["solution"] = df.apply(
            lambda row: f"{row['solution']} 請參閱: {row['see_also']}" if pd.notna(row["see_also"]) and row["see_also"].strip() else row["solution"],
            axis=1
        )

        ws_03 = wb.worksheets[2]  # 第3個工作表(修補建議)
        df_03 = df[df['severity'] != "0"].copy()
        df_03["severity"] = pd.Categorical(df_03["severity"], categories=self.severity_order, ordered=True)
        df_03 = df_03.sort_values(by=["severity", "ip"], ascending=[False, True])   # 先host再風險
        self.write_dataframe_to_sheet(ws_03, df_03, start_row=3)

        ws_04 = wb.worksheets[3]  # 第4個工作表(詳細資訊)
        df["severity"] = pd.Categorical(df["severity"], categories=self.severity_order, ordered=True)
        df = df.sort_values(by=["ip", "severity"], ascending=[True, False])   # 先風險再host
        self.write_dataframe_to_sheet(ws_04, df, start_row=3)

        # 儲存 Excel
        wb.save(output_path)
        print(f"Nessus_Excel --- fin")

    @staticmethod
    def _to_df(json_path):
        # 讀取 JSON 檔案
        with open(json_path, "r", encoding="utf-8") as f:
            json_data = json.load(f)

        # 轉回原本的 dict[dict[DataFrame]]
        dfs = {key: pd.DataFrame(records) for key, records in json_data.items()}
        return dfs


def main():
    # pd.set_option('display.max_columns', None)  # 顯示所有列
    project_path = os.path.join(os.getcwd(), "data/uploadproject", i.taskname)
    

    parser = NessusParser()
    df_nessus, df_nessus_scan_info = parser.nessus_to_df(project_path)

    report_generator = NessusExcelGenerater()
    report_generator.generate_report(df_vuln, project_path)

    # print("報告已生成，總共掃描的漏洞嚴重程度統計：", count_severity)

if __name__ == "__main__":
    main()
    
